import openpyxl
import datetime
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import email

from json_action import *


id_cell_mail = []
id_cell_poste = []
mail_envoie = []
poste = []

current_date = date.today()  
delai = datetime.timedelta(7) # 7 Jours de délai avant envoie du mail
relance_date = (current_date - delai) # La date de la relance

# ---------------------------------------------------------------

def verif_date():
  global ENVOI

  fichier = open("assets/date.txt", 'r')
  if str(date.today()) in fichier.read():
    fichier.close()
    ENVOI = False
  else: 
    ENVOI = True
    ecrire_date()


def ecrire_date():
  global relance_date

  fichier2 = open("assets/date.txt", 'a')
  fichier2.write("\n Relance execute le " + str(date.today()) + " pour 7j avant")

# ---------------------------------------------------------------

def excel_find():
  workbook = openpyxl.load_workbook('assets/candidatures.xlsx', read_only = True)
  sheet = workbook.active

  for row in sheet.iter_rows(min_row = 1, max_row = 1000, min_col = 0, max_col = 10):
    for cell in row:
      
      if str(cell.value) == str(relance_date) + " 00:00:00":
        id_cell_mail.append((cell.row, cell.column + 3))
        id_cell_poste.append((cell.row, cell.column -1))
        

  for i in range(len(id_cell_mail)):

    if sheet.cell(row=id_cell_mail[i][0], column=id_cell_mail[i][1]).value == None:
      pass
    else : 
      mail_envoie.append(sheet.cell(row=id_cell_mail[i][0], column=id_cell_mail[i][1]).value)
      poste.append(sheet.cell(row=id_cell_poste[i][0], column=id_cell_poste[i][1]).value)

  workbook.close()

# ---------------------------------------------------------------

def envoie_mail():

# Custom message
# Ne marche pas avec l'executable mais en lancant le script
# Pour choisir ce mode, dans login.json, passer la valeur de mode à 2

  message2option = {
  "core_message" : f"Madame, Monsieur,\n\n\
Pour faire suite à ma candidature envoyée le {relance_date} pour le poste de {poste[x]} je me permets de revenir vers vous pour savoir qu'elle est l’avancée du processus de recrutement.\n\
Je suis toujours très intéressé par le poste de {poste[x]} au sein de votre entreprise, qui correspond à mes compétences en développement informatique et à mes ambitions professionnelles.\n\
Pour avoir un aperçu de mon travail, voici le lien vers mon github : [Votre lien Github]\n\n\
Je reste à votre entière disposition pour convenir d’un rendez-vous afin de vous faire part de ma motivation et de mes capacités pour le poste de {poste[x]}.\n\n\
Je vous prie d’agréer, Madame, Monsieur, mes salutations distinguées.\n\n\
[Prénom Nom]\n\
[Votre numéro téléphone]\n\
[Votre lien Linkedin] \n"
}

  message_file = open('assets/relance_message.txt', 'r', encoding = 'utf-8')
  message_data = message_file.read()
      
  msg = MIMEMultipart()
  msg['From'] = email
  msg['To'] = f"{mail_envoie[x]}"

  msg['Subject'] = f"Relance candidature au poste de {poste[x]}"

  if mode != 1:
    msg.attach(MIMEText(message2option))
  else:  
    msg.attach(MIMEText(message_data))

  mailserver = smtplib.SMTP('smtp.gmail.com', 587)
  mailserver.ehlo()
  mailserver.starttls()
  mailserver.ehlo()
  mailserver.login(email, password)
  mailserver.sendmail(email, f"{mail_envoie[x]}", msg.as_string())
  mailserver.quit()


# ---------------------------------------------------------------

def number_send():
  global x, ENVOI
  if ENVOI == True:
    for x in range(len(mail_envoie)):
      envoie_mail()
      
    print(f"\n \n {len(mail_envoie)} candidature(s) a bien été envoyé")
    input()


